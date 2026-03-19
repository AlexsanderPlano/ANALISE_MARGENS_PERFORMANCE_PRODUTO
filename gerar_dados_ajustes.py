"""
gerar_dados_ajustes.py — v2.0
Calcula dados de ajuste de estoque usando APENAS fontes dentro da pasta do projeto.
Sem dependencia de logs externos.

Fontes:
  - dados/movimentos/*.xlsx        (movimentos DEPOIS do ajuste)
  - dados/movimentos_antes/*.xlsx  (movimentos ANTES do ajuste)
  - dados/VENDAS_2025.xlsx         (receita por produto/mes)
  - dados/DEVOLUCOES_2025.xlsx     (devolucoes por produto)

Saida:
  - dados-ajustes.js               (array JS para o frontend)
"""

import os
import glob
import openpyxl
from collections import defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))
MOV_DEPOIS = os.path.join(BASE, 'dados', 'movimentos')
MOV_ANTES = os.path.join(BASE, 'dados', 'movimentos_antes')
VENDAS_FILE = os.path.join(BASE, 'dados', 'VENDAS_2025.xlsx')
DEVOL_FILE = os.path.join(BASE, 'dados', 'DEVOLUCOES_2025.xlsx')
OUTPUT_FILE = os.path.join(BASE, 'dados-ajustes.js')

MESES_MAP = {
    'Janeiro':1,'Fevereiro':2,'Março':3,'Abril':4,'Maio':5,'Junho':6,
    'Julho':7,'Agosto':8,'Setembro':9,'Outubro':10,'Novembro':11,'Dezembro':12
}


def processar_movimentos(filepath):
    """Extrai dados de um arquivo de movimentos xlsx.
    Retorna dict com: cmc_mensal, qtd_venda_mensal, cmv_mensal,
    saldo_final, cmc_final, vlr_estoque, ops, diagnostico diario."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    if not rows:
        return None

    # Ultimo CMC de cada mes
    cmc_mes = {}
    qtd_venda_mes = defaultdict(float)
    ops = set()

    # Diagnostico diario: ultimo saldo e CMC de cada dia
    saldos_dia = {}   # {data_str: saldo}
    cmc_dia = {}      # {data_str: cmc}

    for r in rows:
        if not r[0]:
            continue
        try:
            m = r[0].month
            dia_str = r[0].strftime('%Y-%m-%d')
        except AttributeError:
            continue

        if r[5] is not None:
            cmc_mes[m] = float(r[5])
            cmc_dia[dia_str] = abs(float(r[5]))

        if r[4] is not None:
            saldos_dia[dia_str] = float(r[4])

        origem = str(r[1] or '')
        if 'Venda' in origem:
            qtd_venda_mes[m] += abs(float(r[2] or 0))
        if 'Ordem' in origem and 'Produ' in origem and r[8]:
            ops.add(str(r[8]))

    # Diagnostico: contar dias por tipo
    dias_total = len(saldos_dia)
    dias_negativo = sum(1 for s in saldos_dia.values() if s < 0)

    # "Excessivo" = saldo > 3x a media de venda diaria
    qtd_vendida_total = sum(qtd_venda_mes.values())
    # Dias uteis no ano ~252, mas usamos dias com movimento real
    dias_com_venda = sum(1 for r in rows if r[1] and 'Venda' in str(r[1]))
    media_venda_dia = qtd_vendida_total / max(dias_total, 1)
    limiar_excessivo = media_venda_dia * 3
    if limiar_excessivo < 5:
        limiar_excessivo = 40  # fallback para produtos com pouca venda

    dias_excessivo = sum(1 for s in saldos_dia.values() if s > limiar_excessivo)
    dias_ok = dias_total - dias_negativo - dias_excessivo
    pior_saldo = min(saldos_dia.values()) if saldos_dia else 0
    maior_saldo = max(saldos_dia.values()) if saldos_dia else 0

    # CMC: calcular mediana para detectar dias "acima do normal"
    cmc_valores = [v for v in cmc_dia.values() if v > 0]
    if cmc_valores:
        cmc_valores_sorted = sorted(cmc_valores)
        n = len(cmc_valores_sorted)
        cmc_mediana = cmc_valores_sorted[n // 2]
        # "Acima do normal" = CMC > 3x a mediana
        limiar_cmc = cmc_mediana * 3
        dias_cmc_alto = sum(1 for v in cmc_dia.values() if v > limiar_cmc)
        cmc_max = max(cmc_valores)
    else:
        cmc_mediana = 0
        dias_cmc_alto = 0
        cmc_max = 0
        limiar_cmc = 0

    # Saldo e valor estoque final
    last = rows[-1]
    saldo_final = float(last[4]) if last[4] is not None else 0
    cmc_final = float(last[5]) if last[5] is not None else 0
    vlr_estoque = float(last[6]) if last[6] is not None else 0

    # CMV mensal = qtd vendida no mes * ultimo CMC do mes
    cmv_mes = {}
    for m in range(1, 13):
        q = qtd_venda_mes.get(m, 0)
        c = cmc_mes.get(m, 0)
        cmv_mes[m] = q * c

    # Qtd vendida total
    qtd_total = sum(qtd_venda_mes.values())

    return {
        'cmc_mes': cmc_mes,
        'qtd_venda_mes': dict(qtd_venda_mes),
        'cmv_mes': cmv_mes,
        'saldo_final': saldo_final,
        'cmc_final': cmc_final,
        'vlr_estoque': vlr_estoque,
        'ops': len(ops),
        'qtd_total': qtd_total,
        'cmv_total': sum(cmv_mes.values()),
        # Diagnostico diario
        'dias_total': dias_total,
        'dias_negativo': dias_negativo,
        'dias_excessivo': dias_excessivo,
        'dias_ok': dias_ok,
        'pior_saldo': pior_saldo,
        'maior_saldo': maior_saldo,
        'dias_cmc_alto': dias_cmc_alto,
        'cmc_mediana': cmc_mediana,
        'cmc_max': cmc_max,
        'media_venda_dia': media_venda_dia,
        'limiar_excessivo': limiar_excessivo,
    }


def carregar_vendas():
    """Carrega receita mensal por produto de VENDAS_2025.xlsx."""
    print('  Lendo VENDAS_2025.xlsx...')
    wb = openpyxl.load_workbook(VENDAS_FILE, read_only=True)
    ws = wb.active
    next(ws.iter_rows(max_row=1))  # skip header

    vendas = defaultdict(lambda: defaultdict(float))
    qtd_vendas = defaultdict(lambda: defaultdict(float))

    for r in ws.iter_rows(min_row=2, values_only=True):
        prod = str(r[7] or '').strip().upper()
        qtd = float(r[8] or 0)
        total = float(r[10] or 0)
        mes_str = str(r[13] or '')
        m = MESES_MAP.get(mes_str, 0)
        if m and prod:
            vendas[prod][m] += total
            qtd_vendas[prod][m] += qtd

    wb.close()
    print(f'    {len(vendas)} produtos com vendas')
    return vendas, qtd_vendas


def carregar_devolucoes():
    """Carrega devolucoes por produto de DEVOLUCOES_2025.xlsx."""
    print('  Lendo DEVOLUCOES_2025.xlsx...')
    wb = openpyxl.load_workbook(DEVOL_FILE, read_only=True)
    ws = wb.active
    next(ws.iter_rows(max_row=1))  # skip header

    devol_r = defaultdict(float)
    devol_q = defaultdict(float)

    for r in ws.iter_rows(min_row=2, values_only=True):
        prod = str(r[4] or '').strip().upper()
        qtd = abs(float(r[5] or 0))
        total = abs(float(r[7] or 0))
        if prod:
            devol_r[prod] += total
            devol_q[prod] += qtd

    wb.close()
    print(f'    {len(devol_r)} produtos com devolucoes')
    return devol_r, devol_q


def nome_arquivo_para_produto(filename):
    """Converte nome de arquivo para nome de produto.
    Ex: ALHO_PORO_250G.xlsx -> ALHO PORO 250G"""
    return os.path.basename(filename).replace('.xlsx', '').replace('_', ' ')


def cap_cmc_antes(cmc_antes_mes, cmc_depois_mes):
    """Limita CMC 'antes' ao dobro do max CMC 'depois' do mesmo produto.
    Isso corrige distorcoes de saldo negativo sem inventar valores."""
    if not cmc_depois_mes:
        return cmc_antes_mes

    max_depois = max(cmc_depois_mes.values()) if cmc_depois_mes else 50
    cap = max(max_depois * 2, 10)  # minimo R$10

    capped = {}
    for m, v in cmc_antes_mes.items():
        capped[m] = min(abs(v), cap)
    return capped


def main():
    print('=' * 60)
    print('GERAR DADOS AJUSTES v2.0 — Fontes locais apenas')
    print('=' * 60)

    # 1. Carregar vendas e devolucoes
    vendas_mes, qtd_vendas_mes = carregar_vendas()
    devol_r, devol_q = carregar_devolucoes()

    # 2. Listar produtos (intersecao de movimentos antes E depois)
    arqs_depois = {nome_arquivo_para_produto(f): f for f in glob.glob(os.path.join(MOV_DEPOIS, '*.xlsx'))}
    arqs_antes = {nome_arquivo_para_produto(f): f for f in glob.glob(os.path.join(MOV_ANTES, '*.xlsx'))}

    produtos_comuns = sorted(set(arqs_depois.keys()) & set(arqs_antes.keys()))
    print(f'\n  Produtos: {len(arqs_depois)} depois, {len(arqs_antes)} antes, {len(produtos_comuns)} em comum')

    # 3. Processar cada produto
    resultados = []
    erros = 0

    for nome in produtos_comuns:
        try:
            depois = processar_movimentos(arqs_depois[nome])
            antes = processar_movimentos(arqs_antes[nome])

            if not depois or not antes:
                continue

            # Cap CMC antes para evitar distorcoes
            cmc_antes_capped = cap_cmc_antes(antes['cmc_mes'], depois['cmc_mes'])

            # Recalcular CMV antes com CMC capado
            cmv_antes_mes = {}
            for m in range(1, 13):
                q = antes['qtd_venda_mes'].get(m, 0)
                c = cmc_antes_capped.get(m, 0)
                cmv_antes_mes[m] = q * c

            cmv_antes_total = sum(cmv_antes_mes.values())
            cmv_depois_total = depois['cmv_total']

            # Receita (do VENDAS_2025)
            rec_por_mes = vendas_mes.get(nome, {})
            receita = sum(rec_por_mes.values())

            # Devolucoes
            dr = devol_r.get(nome, 0)
            dq = devol_q.get(nome, 0)

            # Qtd vendida (do movimentos depois — fonte da verdade)
            qtd_vendida = depois['qtd_total']

            # Preco medio
            preco_medio = receita / qtd_vendida if qtd_vendida > 0 else 0

            # Quebra %
            quebra = (dq / (qtd_vendida + dq) * 100) if (qtd_vendida + dq) > 0 else 0

            # Margens
            mg_antes = receita - cmv_antes_total
            mg_depois = receita - cmv_depois_total
            pmg_antes = (mg_antes / receita * 100) if receita > 0 else 0
            pmg_depois = (mg_depois / receita * 100) if receita > 0 else 0

            # Economia
            economia = cmv_antes_total - cmv_depois_total

            # Arrays mensais (12 posicoes)
            cmc_mes_antes_arr = [round(cmc_antes_capped.get(m, 0), 2) for m in range(1, 13)]
            cmc_mes_depois_arr = [round(depois['cmc_mes'].get(m, 0), 2) for m in range(1, 13)]
            qtd_mes_arr = [round(depois['qtd_venda_mes'].get(m, 0)) for m in range(1, 13)]
            cmv_mes_antes_arr = [round(cmv_antes_mes.get(m, 0), 2) for m in range(1, 13)]
            cmv_mes_depois_arr = [round(depois['cmv_mes'].get(m, 0), 2) for m in range(1, 13)]
            rec_mes_arr = [round(rec_por_mes.get(m, 0), 2) for m in range(1, 13)]

            resultados.append({
                'nome': nome,
                'qtdVendida': round(qtd_vendida),
                'qtdDevolvida': round(dq),
                'receita': round(receita, 2),
                'devolucao': round(dr, 2),
                'recAntes': round(receita, 2),
                'cmvAntes': round(cmv_antes_total, 2),
                'mgAntes': round(mg_antes, 2),
                'pmgAntes': round(pmg_antes, 2),
                'recDepois': round(receita, 2),
                'cmvDepois': round(cmv_depois_total, 2),
                'mgDepois': round(mg_depois, 2),
                'pmgDepois': round(pmg_depois, 2),
                'saldoAntes': round(antes['saldo_final']),
                'saldoDepois': round(depois['saldo_final']),
                'cmcFinalAntes': round(abs(antes['cmc_final']), 2),
                'cmcFinalDepois': round(abs(depois['cmc_final']), 2),
                'vlrEstoqueAntes': round(antes['vlr_estoque'], 2),
                'vlrEstoqueDepois': round(depois['vlr_estoque'], 2),
                'opsAntes': antes['ops'],
                'opsDepois': depois['ops'],
                'precoMedio': round(preco_medio, 2),
                # Diagnostico diario (ANTES)
                'diasTotal': antes['dias_total'],
                'diasNegativo': antes['dias_negativo'],
                'diasExcessivo': antes['dias_excessivo'],
                'diasOk': antes['dias_ok'],
                'piorSaldo': round(antes['pior_saldo']),
                'maiorSaldo': round(antes['maior_saldo']),
                'diasCmcAlto': antes['dias_cmc_alto'],
                'cmcMediana': round(antes['cmc_mediana'], 2),
                'cmcMax': round(antes['cmc_max'], 2),
                # Diagnostico diario (DEPOIS)
                'diasNegativoDepois': depois['dias_negativo'],
                'diasExcessivoDepois': depois['dias_excessivo'],
                'diasOkDepois': depois['dias_ok'],
                # Arrays mensais
                'cmcMesAntes': cmc_mes_antes_arr,
                'cmcMesDepois': cmc_mes_depois_arr,
                'qtdMes': qtd_mes_arr,
                'cmvMesAntes': cmv_mes_antes_arr,
                'cmvMesDepois': cmv_mes_depois_arr,
                'recMes': rec_mes_arr,
            })

        except Exception as e:
            print(f'  ERRO {nome}: {e}')
            erros += 1

    # Ordenar por nome
    resultados.sort(key=lambda x: x['nome'])

    # 4. Gerar JS
    lines = ['const DADOS_AJUSTES = [']
    for r in resultados:
        lines.append('    {')
        lines.append(f"        nome: '{r['nome']}',")
        lines.append(f"        qtdVendida: {r['qtdVendida']}, qtdDevolvida: {r['qtdDevolvida']}, receita: {r['receita']}, devolucao: {r['devolucao']},")
        lines.append(f"        recAntes: {r['recAntes']}, cmvAntes: {r['cmvAntes']}, mgAntes: {r['mgAntes']}, pmgAntes: {r['pmgAntes']},")
        lines.append(f"        recDepois: {r['recDepois']}, cmvDepois: {r['cmvDepois']}, mgDepois: {r['mgDepois']}, pmgDepois: {r['pmgDepois']},")
        lines.append(f"        saldoAntes: {r['saldoAntes']}, saldoDepois: {r['saldoDepois']},")
        lines.append(f"        cmcFinalAntes: {r['cmcFinalAntes']}, cmcFinalDepois: {r['cmcFinalDepois']},")
        lines.append(f"        vlrEstoqueAntes: {r['vlrEstoqueAntes']}, vlrEstoqueDepois: {r['vlrEstoqueDepois']},")
        lines.append(f"        opsAntes: {r['opsAntes']}, opsDepois: {r['opsDepois']},")
        lines.append(f"        precoMedio: {r['precoMedio']},")
        lines.append(f"        diasTotal: {r['diasTotal']}, diasNegativo: {r['diasNegativo']}, diasExcessivo: {r['diasExcessivo']}, diasOk: {r['diasOk']},")
        lines.append(f"        piorSaldo: {r['piorSaldo']}, maiorSaldo: {r['maiorSaldo']},")
        lines.append(f"        diasCmcAlto: {r['diasCmcAlto']}, cmcMediana: {r['cmcMediana']}, cmcMax: {r['cmcMax']},")
        lines.append(f"        diasNegativoDepois: {r['diasNegativoDepois']}, diasExcessivoDepois: {r['diasExcessivoDepois']}, diasOkDepois: {r['diasOkDepois']},")
        lines.append(f"        cmcMesAntes:  {r['cmcMesAntes']},")
        lines.append(f"        cmcMesDepois: {r['cmcMesDepois']},")
        lines.append(f"        qtdMes:  {r['qtdMes']},")
        lines.append(f"        cmvMesAntes:  {r['cmvMesAntes']},")
        lines.append(f"        cmvMesDepois: {r['cmvMesDepois']},")
        lines.append(f"        recMes: {r['recMes']}")
        lines.append('    },')
    lines.append('];')

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    # 5. Resumo
    tot_rec = sum(r['receita'] for r in resultados)
    tot_cmv_a = sum(r['cmvAntes'] for r in resultados)
    tot_cmv_d = sum(r['cmvDepois'] for r in resultados)
    tot_eco = tot_cmv_a - tot_cmv_d
    mg_a = (tot_rec - tot_cmv_a) / tot_rec * 100 if tot_rec > 0 else 0
    mg_d = (tot_rec - tot_cmv_d) / tot_rec * 100 if tot_rec > 0 else 0

    print(f'\n{"=" * 60}')
    print(f'RESULTADO')
    print(f'{"=" * 60}')
    print(f'  Produtos gerados:   {len(resultados)}')
    print(f'  Erros:              {erros}')
    print(f'  Receita total:      R$ {tot_rec:>14,.2f}')
    print(f'  CMV Antes (capado): R$ {tot_cmv_a:>14,.2f}')
    print(f'  CMV Depois:         R$ {tot_cmv_d:>14,.2f}')
    print(f'  Economia CMV:       R$ {tot_eco:>14,.2f}')
    print(f'  Margem antes:       {mg_a:.1f}%')
    print(f'  Margem depois:      {mg_d:.1f}%')
    print(f'  Variacao:           +{mg_d - mg_a:.1f} p.p.')
    print(f'  Arquivo:            {OUTPUT_FILE}')
    print(f'  Linhas:             {len(lines)}')


if __name__ == '__main__':
    main()
